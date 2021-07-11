# -*- coding: utf-8 -*-
"""
Created on Sat Jul 10 22:02:25 2021

@author: bibek
"""
import openpyxl as xl
import numpy as np
from tkinter import Tk, filedialog     # from tkinter import Tk for Python 3.x
from tkinter.filedialog import askopenfilename
import os
import pandas as pd

Tk().withdraw()
folder_dir = filedialog.askdirectory()

files = os.listdir(folder_dir)
new_dir = os.path.join(folder_dir, 'TEXT_FORMAT')
os.mkdir(new_dir)
wb = xl.Workbook()
for i in files:
    new_name = i.replace('.', '')
    os.rename(folder_dir+'/'+i, new_dir+'/'+new_name+'.txt')

txt_file = os.listdir(new_dir)
m=2
wb.create_sheet('Total')
wb['Total']['A1'] = 'Year'
wb['Total']['B1'] = 'Total'
wb['Total']['C1'] = 'Eror'
 
for i in txt_file:
    year = i[slice(6,10)]
    wb.create_sheet(year)
    dir = os.path.join(new_dir,i)
    data = pd.read_fwf(dir, header=None)
    sheet = wb[year]
    sum = 0
    error = 0
    for j in range(1,366):
        try:
            val = float(data[1][j-1])      
            sheet['A'+str(j)] = data[0][j-1]
            sheet['B'+str(j)] = val
            sum = sum+val
        except:
            sheet['A'+str(j)] = data[0][j-1]
            sheet['B'+str(j)] = data[1][j-1]
            error=error+1
            
    wb['Total']['A'+str(m)] = float(year)
    wb['Total']['B'+str(m)] = sum
    wb['Total']['C'+str(m)] = error
    m=m+1
    
    
            
wb.save(new_dir+'/bibek-Formatted.xlsx')
        
  
        
            
    
    